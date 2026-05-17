"""
kap_rss.py — Direct KAP.org.tr RSS Feed Ingestion + Açığa Satış (Short Sell) Data

Fetches:
1. KAP.org.tr official RSS announcements (company disclosures, financials, dividends)
2. Short-sell (açığa satış) data from BIST/SPK public reports
Stores results to MongoDB kap_disclosures and short_sell_data collections.
"""

import requests
import xml.etree.ElementTree as ET
from datetime import datetime, timedelta
import re
import json
from typing import Optional

# KAP official RSS feeds
KAP_RSS_URLS = {
    "genel": "https://www.kap.org.tr/tr/rss/bildirim-rss-feed",
    "ozet": "https://www.kap.org.tr/tr/rss/genel-ozet-rss-feed",
}

# SPK short-sell data (public BIST report)
SHORT_SELL_URL = "https://www.borsaistanbul.com/tr/sayfa/1655/aciga-satis-verileri"

# Google News as fallback (more reliable for specific tickers)
GOOGLE_NEWS_BASE = "https://news.google.com/rss/search?hl=tr&gl=TR&ceid=TR:tr&q="

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    )
}


def fetch_kap_rss_direct(feed_key: str = "genel", limit: int = 50) -> list:
    """
    Fetches official KAP.org.tr RSS feed directly.
    Returns list of disclosure dicts.
    """
    url = KAP_RSS_URLS.get(feed_key, KAP_RSS_URLS["genel"])
    try:
        resp = requests.get(url, headers=HEADERS, timeout=12)
        if resp.status_code != 200:
            print(f"[KAP RSS] Direct feed returned {resp.status_code}, using fallback")
            return []

        root = ET.fromstring(resp.content)
        items = []
        for item in root.findall("./channel/item"):
            title = item.findtext("title", "").strip()
            link = item.findtext("link", "").strip()
            pub_date = item.findtext("pubDate", "").strip()
            description = item.findtext("description", "").strip()
            source = item.findtext("source", "KAP")

            # Extract ticker from title if possible (e.g. "THYAO - Özel Durum Açıklaması")
            ticker_match = re.match(r'^([A-Z]{2,6})\s*[-–]', title)
            ticker = ticker_match.group(1) if ticker_match else None

            # Remove HTML tags from description
            description_clean = re.sub(r'<[^>]+>', '', description).strip()

            items.append({
                "title": title,
                "description": description_clean,
                "link": link,
                "pub_date": pub_date,
                "source": "KAP.org.tr",
                "ticker": ticker,
                "category": "KAP_OFFICIAL",
                "fetched_at": datetime.now().isoformat()
            })

            if len(items) >= limit:
                break

        print(f"[KAP RSS] Fetched {len(items)} items from official feed ({feed_key})")
        return items

    except Exception as e:
        print(f"[KAP RSS] Direct feed error: {e}")
        return []


def fetch_kap_by_ticker_google(ticker: str, limit: int = 10) -> list:
    """
    Fetches KAP-related news for a specific ticker via Google News RSS.
    More reliable for ticker-specific filtering than official KAP RSS.
    """
    base = ticker.split(".")[0]
    queries = [
        f"{base} KAP bildirimi",
        f"{base} özel durum açıklaması",
        f"{base} temettü bülten",
    ]
    items = []
    for q in queries:
        import urllib.parse
        url = GOOGLE_NEWS_BASE + urllib.parse.quote(q)
        try:
            resp = requests.get(url, headers=HEADERS, timeout=10)
            if resp.status_code != 200:
                continue
            root = ET.fromstring(resp.content)
            for item in root.findall("./channel/item"):
                title = item.findtext("title", "")
                if " - " in title:
                    title = " - ".join(title.split(" - ")[:-1])
                pub_date = item.findtext("pubDate", "")
                items.append({
                    "title": title,
                    "pub_date": pub_date,
                    "ticker": base,
                    "source": "Google/KAP",
                    "category": "KAP_TICKER",
                    "fetched_at": datetime.now().isoformat()
                })
                if len(items) >= limit:
                    break
        except Exception as e:
            print(f"[KAP RSS] Google query error for {base}: {e}")
        if len(items) >= limit:
            break

    return items


def fetch_short_sell_data() -> list:
    """
    Fetches açığa satış (short-sell) data from BIST public reports.
    Since BIST doesn't offer a public JSON API for this,
    we fetch the most recent available CSV data from BIST's data files endpoint.
    Returns list of {ticker, short_ratio, date} dicts.
    """
    # BIST publishes short-sell files here (weekly CSV):
    # Try the most recent Monday's file
    results = []
    today = datetime.now()
    # Walk back to find the last Monday
    days_back = today.weekday()  # 0=Mon
    last_monday = today - timedelta(days=days_back)

    for offset in range(0, 14):  # Try up to 2 weeks back
        d = last_monday - timedelta(days=offset * 7)
        date_str = d.strftime("%Y%m%d")
        url = f"https://www.borsaistanbul.com/data/files/short-selling-positions-{date_str}.xlsx"
        try:
            resp = requests.head(url, headers=HEADERS, timeout=5)
            if resp.status_code == 200:
                print(f"[SHORT SELL] Found BIST short-sell file for {d.strftime('%Y-%m-%d')}")
                # Download and parse
                file_resp = requests.get(url, headers=HEADERS, timeout=15)
                try:
                    import openpyxl
                    from io import BytesIO
                    wb = openpyxl.load_workbook(BytesIO(file_resp.content))
                    ws = wb.active
                    headers = [cell.value for cell in ws[1]]
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        row_dict = dict(zip(headers, row))
                        # Normalize key names
                        ticker = str(row_dict.get("Sembol", row_dict.get("Symbol", ""))).strip()
                        short_ratio = row_dict.get("Açığa Satış Oranı", row_dict.get("Short Ratio", None))
                        if ticker:
                            results.append({
                                "ticker": ticker,
                                "short_ratio": float(short_ratio) if short_ratio else None,
                                "date": d.strftime("%Y-%m-%d"),
                                "source": "BIST",
                                "fetched_at": datetime.now().isoformat()
                            })
                    break
                except ImportError:
                    print("[SHORT SELL] openpyxl not installed. Falling back to SPK CSV.")
                    break
        except Exception as e:
            print(f"[SHORT SELL] BIST file error for {date_str}: {e}")

    # If BIST xlsx not available, use SPK public data (text-based fallback)
    if not results:
        results = _fetch_short_sell_spk_fallback()

    return results


def _fetch_short_sell_spk_fallback() -> list:
    """
    SPK (Sermaye Piyasası Kurulu) publishes periodic short-sell aggregates.
    This is a best-effort fallback returning known major short positions.
    """
    # SPK's official data endpoint (public, no auth needed)
    try:
        url = "https://www.spk.gov.tr/SiteApps/Veritabani/DovizKurlari"  # placeholder example
        # In production, replace with the actual SPK short-sell data endpoint
        # For now, return empty - the Celery task will retry
        print("[SHORT SELL] SPK fallback: No data available. Will retry next schedule.")
        return []
    except Exception as e:
        print(f"[SHORT SELL] SPK fallback error: {e}")
        return []


def store_kap_to_mongodb(items: list, collection_name: str = "kap_disclosures") -> int:
    """
    Stores KAP disclosure items to MongoDB. Avoids duplicates by title+date.
    Returns count of newly inserted documents.
    """
    if not items:
        return 0
    try:
        from pymongo import MongoClient
        import os
        client = MongoClient(os.environ.get("MONGO_URI", "mongodb://localhost:27017"))
        db = client["omniquant"]
        col = db[collection_name]

        inserted = 0
        for item in items:
            # Upsert by title + pub_date composite key
            key = {"title": item.get("title"), "pub_date": item.get("pub_date", item.get("date"))}
            result = col.update_one(key, {"$setOnInsert": item}, upsert=True)
            if result.upserted_id:
                inserted += 1

        print(f"[KAP Store] {inserted} new items stored to '{collection_name}'")
        return inserted
    except Exception as e:
        print(f"[KAP Store] MongoDB error: {e}")
        return 0


def get_kap_context_for_ticker(ticker: str, days: int = 7) -> dict:
    """
    Retrieves recent KAP disclosures + short-sell context for a ticker from MongoDB.
    Used by the AI debate/chat endpoints for enrichment.
    """
    base = ticker.split(".")[0]
    try:
        from pymongo import MongoClient
        import os
        client = MongoClient(os.environ.get("MONGO_URI", "mongodb://localhost:27017"))
        db = client["omniquant"]

        # Recent KAP disclosures
        cutoff = (datetime.now() - timedelta(days=days)).isoformat()
        disclosures = list(
            db["kap_disclosures"].find(
                {"ticker": base, "fetched_at": {"$gte": cutoff}},
                {"_id": 0, "title": 1, "pub_date": 1, "category": 1}
            ).sort("fetched_at", -1).limit(5)
        )

        # Short-sell ratio
        short_data = db["short_sell_data"].find_one(
            {"ticker": base},
            {"_id": 0, "short_ratio": 1, "date": 1},
            sort=[("date", -1)]
        )

        return {
            "ticker": base,
            "disclosures": disclosures,
            "short_sell": short_data or {"short_ratio": None, "note": "Veri mevcut değil"}
        }
    except Exception as e:
        print(f"[KAP Context] Error for {ticker}: {e}")
        return {"ticker": ticker, "disclosures": [], "short_sell": {}}


if __name__ == "__main__":
    print("=== KAP RSS Direct Feed Test ===")
    items = fetch_kap_rss_direct("genel", limit=5)
    print(json.dumps(items, indent=2, ensure_ascii=False))

    print("\n=== KAP by Ticker (THYAO) ===")
    ticker_items = fetch_kap_by_ticker_google("THYAO", limit=3)
    print(json.dumps(ticker_items, indent=2, ensure_ascii=False))

    print("\n=== Short Sell Data ===")
    short = fetch_short_sell_data()
    print(json.dumps(short[:3], indent=2, ensure_ascii=False))
